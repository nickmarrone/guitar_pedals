import sys
import re
import pandas as pd
from pathlib import Path
from natsort import natsorted
import openpyxl
from openpyxl.utils import get_column_letter
import argparse
from openpyxl.styles import PatternFill

EXCLUDED_KEYWORDS = ["ic socket", "enclosure", "dust cover"]

SORT_ORDER = {
    "Resistor": 0,
    "Capacitor": 1,
    "Diode": 2,
    "Transistor": 3,
    "IC": 4,
    "Potentiometer": 5,
    "Switch": 6,
    "Connector": 7,
    "Other": 8,
}

def extract_relevant_columns(df):
    # Handle different column structures in original files
    available_columns = df.columns.tolist()
    
    # Map possible column names to our standard names
    column_mapping = {
        "Part": ["Part"],
        "Value": ["Value"],
        "Description": ["Description"],
        "Notes": ["Notes", "Link URL (non-Mouser)"]
    }
    
    # Preserve existing Type column if it exists
    cols_to_extract = []
    if "Type" in available_columns:
        cols_to_extract.append("Type")
    
    # Add standard columns that exist
    for standard_name, possible_names in column_mapping.items():
        for possible_name in possible_names:
            if possible_name in available_columns:
                cols_to_extract.append(possible_name)
                break
    
    return df[cols_to_extract]

def normalize_notes(notes):
    return notes.fillna("").str.strip()

def decode_euro_decimal(val: str) -> str:
    """
    Converts resistor and capacitor values like 2K2 or 6n8 to 2.2K or 6.8n
    """
    if not isinstance(val, str):
        return val
    match = re.match(r'^(\d+)([RrKkMmNnUuPp])(\d+)$', val.replace(" ", ""))
    if match:
        return f"{match.group(1)}.{match.group(3)}{match.group(2)}".lower()
    return val.lower().strip()

def convert_to_euro_notation(val: str) -> str:
    """
    Converts capacitor values to Euro-style notation to match inventory format.
    Examples: "1.5n" -> "1n5", "2.2n" -> "2n2", "100n" -> "100n", "47p" -> "47p"
    Also removes 'f' suffix: "47pf" -> "47p", "100nf" -> "100n"
    """
    if not isinstance(val, str):
        return val
    
    val = val.strip().lower()
    
    # Remove 'f' suffix if present
    val = re.sub(r'f$', '', val)
    
    # If already in Euro-style notation, return as is
    if re.match(r'^\d+[pnu]\d*$', val):
        return val
    
    # Extract numeric value and unit
    match = re.match(r'([\d.]+)([pnu])', val)
    if match:
        num = float(match.group(1))
        unit = match.group(2)
        
        # Convert to Euro-style notation
        if num == int(num):
            # Whole number: "100n" -> "100n"
            return f"{int(num)}{unit}"
        else:
            # Decimal: "1.5n" -> "1n5", "2.2n" -> "2n2"
            whole_part = int(num)
            decimal_part = int((num - whole_part) * 10)
            return f"{whole_part}{unit}{decimal_part}"
    
    # If no unit found, return as is
    return val

def parse_resistor_value(value_str):
    try:
        value_str = decode_euro_decimal(value_str).upper()
        if "M" in value_str:
            return float(value_str.replace("M", "")) * 1e6
        elif "K" in value_str:
            return float(value_str.replace("K", "")) * 1e3
        elif "R" in value_str:
            return float(value_str.replace("R", ""))
        else:
            return float(value_str)
    except:
        return float('inf')

def parse_capacitor_value(value_str):
    """
    Returns a tuple: (unit_rank, numeric_value)
    Where unit_rank is an int representing: pF=0, nF=1, uF=2
    """
    unit_order = {'p': 0, 'n': 1, 'u': 2}
    try:
        val = decode_euro_decimal(value_str).lower().replace(" ", "")
        match = re.match(r'([\d.]+)([pnu])f?', val)
        if match:
            num = float(match.group(1))
            unit = match.group(2)
            return (unit_order.get(unit, 99), num)
        else:
            return (99, float(val))  # fallback: numeric only
    except:
        return (99, float('inf'))


def get_type(description, value=None, part=None):
    if not isinstance(description, str):
        return "Other"
    d = description.lower()
    
    # Check for LED (diode)
    if "led" in d:
        return "Diode"
    
    # Check for IC part numbers (IC followed by digits)
    if part and isinstance(part, str) and re.match(r'^IC\d+$', part.strip(), re.IGNORECASE):
        return "IC"
    
    # Check for J-number transistors in value field
    if value and isinstance(value, str) and re.match(r'^J\d+$', value.strip(), re.IGNORECASE):
        return "Transistor"
    
    if "resistor" in d:
        return "Resistor"
    elif "capacitor" in d:
        return "Capacitor"
    elif "diode" in d:
        return "Diode"
    elif "transistor" in d:
        return "Transistor"
    elif "ic" in d or "opamp" in d or "operational amplifier" in d:
        return "IC"
    elif re.search(r'\bpot\b', d) or "trimmer" in d:
        return "Potentiometer"
    elif re.search(r'\bswitch\b', d):
        return "Switch"
    elif "jack" in d or "connector" in d:
        return "Connector"
    else:
        return "Other"

def description_is_excluded(desc):
    if not isinstance(desc, str):
        return False
    desc = desc.lower()
    return any(term in desc for term in EXCLUDED_KEYWORDS)

def sort_bom(df):
    # Check if Part column exists, otherwise use None
    if "Part" in df.columns:
        df["Type"] = df.apply(lambda row: get_type(row["Description"], row["Value"], row["Part"]), axis=1)
    else:
        df["Type"] = df.apply(lambda row: get_type(row["Description"], row["Value"], None), axis=1)

    def sort_key(row):
        typ = row["Type"]
        order = SORT_ORDER.get(typ, 99)
        val = row["Value"]
        if typ == "Resistor":
            parsed_val = parse_resistor_value(val)
        elif typ == "Capacitor":
            parsed_val = parse_capacitor_value(val)
        else:
            parsed_val = str(val).lower()
        return (order, parsed_val)

    df["SortKey"] = df.apply(sort_key, axis=1)
    df = df.sort_values(by="SortKey")
    df = df.drop(columns=["SortKey"])
    return df

def sort_combined_bom(df):
    """
    Sort combined BOM without recalculating Type (since it already exists)
    """
    def sort_key(row):
        typ = row["Type"]
        order = SORT_ORDER.get(typ, 99)
        val = row["Value"]
        if typ == "Resistor":
            parsed_val = parse_resistor_value(val)
        elif typ == "Capacitor":
            parsed_val = parse_capacitor_value(val)
        else:
            parsed_val = str(val).lower()
        return (order, parsed_val)

    df["SortKey"] = df.apply(sort_key, axis=1)
    df = df.sort_values(by="SortKey")
    df = df.drop(columns=["SortKey"])
    return df

def process_bom_file(file_path):
    xl = pd.ExcelFile(file_path)
    sheet_names = xl.sheet_names

    # Process all sheets except the first one (Instructions) and any combined sheets
    # Look for sheets that contain parts data
    relevant_sheets = []
    for sheet in sheet_names:
        if "instruction" not in sheet.lower() and "combined" not in sheet.lower():
            relevant_sheets.append(sheet)
    
    combined_df = pd.DataFrame()

    for sheet in relevant_sheets:
        df = xl.parse(sheet)
        df = extract_relevant_columns(df)
        combined_df = pd.concat([combined_df, df], ignore_index=True)

    # Filter out excluded descriptions
    combined_df = combined_df[~combined_df["Description"].apply(description_is_excluded)]
    
    # Filter out rows with missing or invalid essential data
    combined_df = combined_df.dropna(subset=["Description", "Value"])
    combined_df = combined_df[combined_df["Description"].str.strip() != ""]
    combined_df = combined_df[combined_df["Value"].str.strip() != ""]

    combined_df["Part"] = combined_df["Part"].astype(str)
    combined_df["Notes"] = normalize_notes(combined_df["Notes"])
    
    # Calculate Type BEFORE grouping
    combined_df["Type"] = combined_df.apply(lambda row: get_type(row["Description"], row["Value"], row["Part"]), axis=1)

    # Group by Value + Description + Notes + Type
    grouped = (
        combined_df.groupby(["Value", "Description", "Notes", "Type"], dropna=False)
        .agg({
            "Part": lambda x: ", ".join(natsorted(set(x.dropna()))),
        })
        .reset_index()
    )

    grouped = grouped[["Type", "Part", "Value", "Description", "Notes"]]
    
    # Sort the grouped data
    grouped = sort_combined_bom(grouped)
    
    return grouped

def get_aion_fx_name(file_name):
    return Path(file_name).stem.split(" - ")[0]

def count_parts_in_row(part_str):
    """
    Count the number of parts in a comma-separated part string.
    For example: "Q1, Q2, Q3, Q4, Q5" returns 5
    """
    if pd.isna(part_str) or not part_str:
        return 0
    # Split by comma and count non-empty parts
    parts = [p.strip() for p in str(part_str).split(",") if p.strip()]
    return len(parts)

def autofit_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column index (1-based)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

def load_resistor_inventory_from_xlsx(filepath):
    df = pd.read_excel(filepath, sheet_name="TH Resistors", usecols=[0, 1], header=None)
    inventory = {}
    for _, row in df.iterrows():
        val = str(row[0]).strip().lower()
        if not val:
            continue
        status = str(row[1]).strip().lower() if len(row) > 1 else ""
        inventory[val] = status
    return inventory

def interpret_inventory_amount(raw):
    if pd.isna(raw):
        return "ok"  # NaN/empty amount = plenty in stock
    val = str(raw).strip().lower()
    if val == "few":
        return "few"  # "few" = low stock
    try:
        if float(val) > 0:
            return "ok"  # positive number = plenty in stock
        else:
            return "ok"  # zero or negative = still available
    except ValueError:
        return "ok"  # non-numeric = assume available

def load_capacitor_inventory_from_xlsx(filepath):
    df = pd.read_excel(filepath, sheet_name="TH Capacitors", header=0)

    inventory = {
        "ceramic": {},
        "film": {},
        "electrolytic": {},
    }

    for _, row in df.iterrows():
        # Ceramic: columns 0 and 1
        val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if val:
            key = val.lower()  # Keep Euro-style notation like "10p", "22p"
            status = interpret_inventory_amount(row.iloc[1]) if len(row) > 1 else None
            inventory["ceramic"][key] = status

        # Film: columns 3 and 4
        val = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ""
        if val:
            key = val.lower()  # Keep Euro-style notation like "1n", "1n5", "2n2"
            status = interpret_inventory_amount(row.iloc[4]) if len(row) > 4 else None
            inventory["film"][key] = status

        # Electrolytic: columns 5 and 6
        val = str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else ""
        if val:
            key = val.lower()  # Keep Euro-style notation
            status = interpret_inventory_amount(row.iloc[6]) if len(row) > 6 else None
            inventory["electrolytic"][key] = status

    print("\n🔍 Debug: Capacitor inventory loaded:")
    for cap_type in ["ceramic", "film", "electrolytic"]:
        print(f"  {cap_type}: {len(inventory[cap_type])} items")
        for key, status in sorted(inventory[cap_type].items()):  # Show all items
            print(f"    '{key}' -> {status}")

    return inventory

def highlight_missing_parts(ws, resistor_inv, capacitor_inv):
    pink_fill = PatternFill(start_color="ffc0cb", end_color="ffc0cb", fill_type="solid")   # missing
    orange_fill = PatternFill(start_color="ffd8a8", end_color="ffd8a8", fill_type="solid") # few

    for row in range(2, ws.max_row + 1):
        value = str(ws.cell(row=row, column=3).value).strip().lower()
        desc = str(ws.cell(row=row, column=4).value or "").lower()

        highlight = None

        if "resistor" in desc:
            status = resistor_inv.get(value)
            if status is None:
                highlight = pink_fill
            elif status == "few":
                highlight = orange_fill

        elif "capacitor" in desc:
            desc_clean = desc.strip().lower()

            cap_type = "other"
            if "electrolytic" in desc_clean:
                cap_type = "electrolytic"
            elif "ceramic" in desc_clean or "mlcc" in desc_clean:
                cap_type = "ceramic"
            elif "film" in desc_clean:
                cap_type = "film"
            elif "tantalum" in desc_clean:
                cap_type = "electrolytic"  # Treat tantalum as electrolytic for inventory purposes

            if cap_type in capacitor_inv:
                # Convert BOM value to Euro-style notation to match inventory
                cap_value = convert_to_euro_notation(value.strip().lower())
                status = capacitor_inv.get(cap_type, {}).get(cap_value)
                print(f"🔎 Checking {cap_type} capacitor '{cap_value}' (from '{value.strip().lower()}') ... status: {status}")
                
                if status is None:
                    highlight = pink_fill
                    print(f"    -> Marking as missing (status is None)")
                elif status == "few":
                    highlight = orange_fill
                    print(f"    -> Marking as few (status is 'few')")
                else:
                    print(f"    -> Available (status is '{status}')")
            else:
                highlight = pink_fill
                print(f"🔎 Unknown capacitor type '{cap_type}' for '{desc_clean}' -> marking as missing")

        if highlight:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = highlight

def main():
    parser = argparse.ArgumentParser(description="Merge Aion FX BOMs into one Excel file.")
    parser.add_argument('--in', nargs='+', dest='input_files', required=True, help='Input BOM .xlsx files')
    parser.add_argument('--out', required=True, help='Output Excel filename')
    parser.add_argument('--inventory', help='Inventory XLSX file with resistor list (optional)')

    args = parser.parse_args()
    input_files = args.input_files
    output_file = args.out
    inventory_file = args.inventory

    valid_sheets = []
    combined_all = pd.DataFrame()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for file_path in input_files:
            aion_fx_name = get_aion_fx_name(file_path)
            bom_df = process_bom_file(file_path)
            if bom_df is not None:
                bom_df.to_excel(writer, sheet_name=aion_fx_name[:31], index=False)
                valid_sheets.append(aion_fx_name[:31])

                df_for_combined = bom_df.copy()
                df_for_combined = df_for_combined[["Type", "Part", "Value", "Description"]]
                df_for_combined["Count"] = df_for_combined["Part"].apply(count_parts_in_row)
                df_for_combined = df_for_combined[["Type", "Value", "Description", "Count"]]
                combined_all = pd.concat([combined_all, df_for_combined], ignore_index=True)
            else:
                print(f"⚠️ Skipped: {file_path}")

        if not valid_sheets:
            print("❌ No valid sheets were created. Exiting without writing file.")
            return

        if not combined_all.empty:
            combined_grouped = (
                combined_all.groupby(["Type", "Value", "Description"], dropna=False)
                .agg({"Count": "sum"})
                .reset_index()
            )
            combined_grouped = combined_grouped[["Type", "Count", "Value", "Description"]]
            combined_sorted = sort_combined_bom(combined_grouped)
            combined_sorted.to_excel(writer, sheet_name="Combined", index=False)
            valid_sheets.append("Combined")

    wb = openpyxl.load_workbook(output_file)
    for sheet_name in valid_sheets:
        ws = wb[sheet_name]
        autofit_column_widths(ws)

        if sheet_name == "Combined" and inventory_file:
            resistor_inventory = load_resistor_inventory_from_xlsx(inventory_file)
            capacitor_inventory = load_capacitor_inventory_from_xlsx(inventory_file)
            highlight_missing_parts(ws, resistor_inventory, capacitor_inventory)

    wb.save(output_file)
    print(f"✅ Output written to {output_file}")

if __name__ == "__main__":
    main()

